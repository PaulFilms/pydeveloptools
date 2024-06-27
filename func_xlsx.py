'''
Toolkit with simplified functions and methods for create .xlsx Reports


`TASK:`
    - BUG(s)
    - si OpenPyXL no se actualiza en un tiempo, hay que plantear el uso de xlsxwriter y xlsxreader
    - Ver actualizaciones de openpyxl para RichText
    - Añadir la funcion de trabajar con letras y enteros:
        ejm: [ COL = xls.utils.get_column_letter(COLUMN) ]
    - Añadir libreria xlsxwriter para uso de hojas EXCEL ya creadas
    - Replantear las fuentes siguiendo el esquema de MARKDOWN (ejm: H1, H2, H3...)
    - wb.create_named_range('_xlnm.Print_Area', ws, 'A1:U56',scope=0) ** Para rangos definidos por nombre

`WARNINGS:`
    - Eliminado de las funciones el self.FONT para incluir el uso de la clase FONTS
    - Se elimina los Try/Except al incio de la clase, en caso de error hay que tratarlo antes

________________________________________________________________________________________________ '''

__update__ = '2024.06.26'
__author__ = 'PABLO GONZALEZ PILA <pablogonzalezpila@gmail.com>'

''' SYSTEM LIBRARIES '''
import os
from enum import Enum
from typing import List

''' IMPORTED LIBRARIES '''
import pandas as pd
import openpyxl as xls
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from openpyxl.worksheet import pagebreak
from openpyxl.utils import get_column_letter


''' OPENPYXL VARIABLES AND FUNCTIONS
________________________________________________________________________________________________ '''

def CELL_STR(ROW: int, COLUMN: int) -> str:
    '''
    Get the Reference String of selected Cell by numbers
    '''
    col_str = xls.utils.get_column_letter(COLUMN)
    cellstr = "{}{}".format(col_str, ROW)
    return cellstr

def COLUMN_STR(COLUMN: int) -> str:
    '''
    Get the Reference Letter of the selected Column by number
    '''
    letter = xls.utils.get_column_letter(COLUMN)
    return letter

class ALIGN_H(Enum):
    '''
    horizontal aligments types
    '''
    GENERAL = "general"
    LEFT = "left"
    CENTER = "center"
    RIGHT = "right"
    FILL = "fill"
    JUSTIFY = "justify"
    CENTER_CONTINUOUS = "centerContinuous"
    DISTRIBUTED = "distributed"

class ALIGN_V(Enum):
    '''
    vertical aligments types
    '''
    TOP = "top"
    CENTER = "center"
    BOTTOM = "bottom"
    JUSTIFY = "justify"
    DISTRIBUTED = "distributed"

class FONTS(Enum):
    '''
    Normalized Fonts
    '''
    TITLE = Font(name='Calibri', size=12, bold=True)
    HEADER = Font(name='Calibri', size=10, bold=True)
    MAIN = Font(name='Calibri', size=10, bold=False)
    CAPTION = Font(name='Calibri', size=8, bold=False)

''' OPENPYXL REPORT
________________________________________________________________________________________________ '''

class XLSREPORT:
    '''
    Sub-Module to make and edit .xlsx Reports

    ### Args:
        - `path` (str): Complete or relative path of report file
        - `WS_NAME` (str): Name of current DataSheet
    
    #### Warnings:
        - This module it's under test yet, Glitches may occur
    '''
    def __init__(self, path: str, worksheet_name: str = "Data") -> None:
        ## PATH
        self.filePath = path
        extension = os.path.splitext(path)[1]
        if extension == str() or extension == None:
            self.filePath += ".xlsx"

        ## INIT
        self.ROW = 1
        self.ALIGN = Alignment(horizontal=ALIGN_H.LEFT.value, vertical=ALIGN_V.CENTER.value)
        # self.worksheet_name = worksheet_name
        
        ## NEW WORKBOOK
        if os.path.exists(self.filePath) == False:
            self.WB = xls.Workbook(self.filePath)
            self.WB.create_sheet(worksheet_name)
            self.WB.save(self.filePath)
            self.WB.close()

        ## LOAD WORKBOOK
        self.WB = xls.load_workbook(self.filePath, read_only=False) ## Force load_workbook to open like WriteOnly
        
        ## WORKSHEET
        if worksheet_name in self.WB.sheetnames:
            self.WS = self.WB[worksheet_name]
        else:
            self.WB.create_sheet(worksheet_name)
            self.WS = self.WB[worksheet_name]
            self.WB.save(self.filePath)
        
        ## FORMAT
        # self.WS.dimensions.ColumnDimension(self.WS, bestFit=True)        

    def SAVE(self) -> None:
        self.WB.save(self.filePath)

    def close(self) -> None:
        self.WB.save(self.filePath)
        self.WB.close()

    def GET_PROPERTIES(self) -> any:
        return self.WB.properties

    def SHEET_NEW(self, sheet_name: str) -> None:
        '''
        Create a new excel sheet
        '''
        self.WB.create_sheet(sheet_name)
        self.WS = self.WB[sheet_name]

    def SHEET_SELECT(self, sheet_name) -> None:
        self.WS = self.WB[sheet_name]
    
    def SHEET_LIST(self) -> List[str]:
        return self.WB.sheetnames 

    ''' FUNCTIONS '''

    def ROW_INC(self, NUMBER: int=1) -> None:
        '''
        Add an increment in row count
        '''
        self.ROW = self.ROW + int(NUMBER)

    def ROW_WIDTH(self, ROW: int, WIDTH: float = 10) -> None:
        '''
        Set height of a row
        '''
        self.WS.row_dimensions[ROW].height = WIDTH

    def COL_WIDTH(self, COL: int, WIDTH: float = 20) -> None:
        '''
        '''
        COL_LETT = COLUMN_STR(COL)
        self.WS.column_dimensions[COL_LETT].width = WIDTH

    def COL_AUTOFIT(self) -> None:
        '''
        Auto-Adjust the Column Width 
        '''
        for column_cells in self.WS.columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (xls.utils.get_column_letter(column_cells[0].column))
                if new_column_length > 0:
                    self.WS.column_dimensions[new_column_letter].width = new_column_length*1.23

    def COL_FILTERS(self) -> None:
        '''
        Set filters in current WorkSheet from A1 to maximun column and maximun row
        '''
        fullRange = f"A1:{get_column_letter(self.WS.max_column)}{self.WS.max_row}"
        self.WS.auto_filter.ref = fullRange

    ''' WRITE FUNCTIONS '''

    def RD(self, ROW: int, COLUMN: int) -> bool | str | int | float:
        '''
        Returns value of selected row and column from current sheet
        '''
        return self.WS.cell(row=ROW, column=COLUMN).value

    def WR(self, ROW: int, COLUMN: int, VALUE = str(), FONT: Font = FONTS.MAIN.value):
        '''
        Type the selected cell in specific formatting
        - `size:` Font Size (10)
        - `bold:` Font Bold (False)
        - `font_name:` Font Name ("Arial")
        '''
        try:
            self.WS.cell(ROW, COLUMN).value = VALUE
            self.WS.cell(ROW, COLUMN).alignment = self.ALIGN
            self.WS.cell(ROW, COLUMN).font = FONT
        except Exception as e:
            print("ERROR WR:")
            print(e)
            self.WS.cell(ROW, COLUMN).value = "ERROR"
    
    def WR_TITLE(self, ROW: int, COLUMN: int, VALUE = ""):
        '''
        Edit selected cell in Title format
        '''
        self.WS.cell(ROW, COLUMN).value = VALUE
        self.WS.cell(ROW, COLUMN).alignment = self.ALIGN
        self.WS.cell(ROW, COLUMN).font = FONTS.TITLE.value
        self.ROW_WIDTH(ROW, 40)

    def WR_HEADER(self, ROW: int, COLUMN: int, VALUE = "", vertical_alignment: str = ALIGN_V.CENTER.value, wrap_text: bool = False):
        '''
        Edit selected cell in Header format
        '''
        self.WS.cell(ROW, COLUMN).value = VALUE
        self.WS.cell(ROW, COLUMN).alignment = Alignment(horizontal='left', vertical=vertical_alignment, wrap_text=wrap_text)
        self.WS.cell(ROW, COLUMN).font = FONTS.HEADER.value

    def WR_HEADERS(self, ROW: int, COLUMN_INIT: int = 1, HEADERS: list = list(), vertical_alignment: str = ALIGN_V.CENTER.value, wrap_text: bool = False):
        '''
        Write and edit format of Headers List
        '''
        for head in HEADERS:
            self.WR_HEADER(ROW=ROW, COLUMN=COLUMN_INIT+HEADERS.index(head), VALUE=head, vertical_alignment=vertical_alignment, wrap_text=wrap_text)
        self.ROW_WIDTH(ROW, 35)

    def LOW_BORDER(self, ROW=1, col_ini=1, col_fin=300):
        '''
        BUG: INCOMPLETE
        Hay que saber bien el diseño y todas las funciones de borders

        https://openpyxl.readthedocs.io/en/stable/styles.html?highlight=border_style
        '''
        # Style = "thick" (Grueso)
        border0 = borders.Side(style = None, color = None, border_style = None)
        borderLow = borders.Side(
            style = "medium", 
            color="000000", 
            # border_style = "double"
            )
        thin = Border(left = border0, right = border0, bottom = borderLow, top = border0)
        for col in range(col_ini, col_fin): 
            self.WS.cell(row=ROW, column=col).border = thin


    ''' UNDER TEST
    ________________________________________________________________________________________________ '''
    
    def WARP(self, ROW: int, COLUMN: int):
        '''
        '''
        self.WS.cell(ROW, COLUMN).alignment = Alignment(wrap_text=True)

    # def MERGE(self, ROW_INI, COL_INI, ROW_FIN, COL_FIN):
    #     self.WS.merge_cells(
    #         start_row = ROW_INI, 
    #         start_column = COL_INI, 
    #         end_row = ROW_FIN, 
    #         end_column = COL_FIN)
    
    def PRNT_AREA(self, COL_FIN: int):
        '''
        Ajusta la zona de impresion
        INCOMPLETE
        '''
        # self.WS = self.WB[SHEET]
        # self.WS.page_setup.orientation = self.WS.ORIENTATION_LANDSCAPE
        # self.WS.page_setup.fitToPage = True
        self.WS.page_setup.fitToPage = 1
        self.WS.page_setup.fitToHeight = False
        COL_STR = COLUMN_STR(COL_FIN)
        self.WS.print_area = "A:" + COL_STR

    def SHEET_HEAD(self, ROW_FIN: int):
        '''
        Define la cabecera superior
        '''
        self.WS.print_title_rows = "1:" + str(ROW_FIN)
        self.WS.page_margins.top = 0.4
        self.WS.page_margins.botom = 0.4
        # self.WS.page_margins.header = 0.7
        self.WS.page_margins.header = 0.4
        self.WS.page_margins.footer = 0.4

    def PAGE_BREAK(self, ROW: int = 1):
        '''
        Insert a page break in selected row
        '''
        page_break = pagebreak.Break(id=ROW-1)
        break_list = self.WS.row_breaks
        break_list.append(page_break)
        self.WS.row_breaks = break_list

    def IMAGE_INSERT(self, ROW: int = 1, COLUMN: int = 1, HEIGHT=None, WIDTH=None):
        '''
        * Necesary install Pillow packages
        INCOMPLETE
        Sin usar ni comprobar
        '''
        PATH = "C:/Users/GONZA_PA/Desktop/rus-logo.png"
        img = Image(PATH)
        ## PIXEL VALUE
        if HEIGHT and WIDTH: 
            img.height = HEIGHT
            img.width = WIDTH
        cell_str = xls.utils.get_column_letter(COLUMN) + str(ROW)
        self.WS.add_image(img, cell_str)


def DF_REPORT(DATAFRAME: pd.DataFrame, fileName: str, fontName: str = 'Calibri') -> None:
    '''
    Create excel report from selected Pandas DataFrame
    '''
    report = XLSREPORT(fileName, fontName)
    ## HEADERS
    HEADERS: list = DATAFRAME.columns.values.tolist()
    report.WR_HEADERS(1, HEADERS, vertical_alignment="top")
    report.COL_FILTERS()
    report.LOW_BORDER(report.ROW, col_fin=len(HEADERS)+1)
    report.ROW_INC()
    ## DATA
    for row in range(len(DATAFRAME.index)):
        row_data = list(DATAFRAME.iloc[row].values)
        for value in row_data:
            report.WR(report.ROW, row_data.index(value)+1, value)
        report.ROW_INC()
    report.COL_AUTOFIT()
    ## FIN
    report.SAVE()


''' UNDER TEST
________________________________________________________________________________________________ '''

...
